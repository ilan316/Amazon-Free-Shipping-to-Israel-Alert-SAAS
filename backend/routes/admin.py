import asyncio
import io
import logging
import os
import uuid
from datetime import datetime, timedelta

logger = logging.getLogger(__name__)
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse, FileResponse
from jose import JWTError, jwt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete, text, or_, and_

from backend.database import get_db, AsyncSessionLocal


# --- Concurrency helpers: run independent statements on their own sessions ---
# The DB sits in a different region (~215ms/round-trip), so collapsing many
# sequential round-trips into a few concurrent ones is the dominant speedup.
async def _one(stmt):
    async with AsyncSessionLocal() as s:
        return (await s.execute(stmt)).one()


async def _scalar(stmt):
    async with AsyncSessionLocal() as s:
        return (await s.execute(stmt)).scalar()


async def _all(stmt):
    async with AsyncSessionLocal() as s:
        return (await s.execute(stmt)).all()
from sqlalchemy import cast, Date
from backend.models import User, Product, UserProduct, NotificationLog, SystemSetting, EmailClick, EmailTemplate, EmailOpen, EmailSendLog, EmailSendRecipient, BlogPublishedAsin, BlogDismissedAsin, BlogDraft, BlogDraftJob, CategoryTranslation
from backend.blog_utils import fetch_amazon_product, generate_with_claude, build_post_html, commit_to_github, publish_draft, add_to_prices_page, get_github_file, delete_github_file, remove_from_prices_page
from backend.scheduler import queue_blog_social_post
from backend.auth import get_current_admin, hash_password, verify_password, SECRET_KEY, ALGORITHM


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class RequestEmailChangeRequest(BaseModel):
    new_email: str
    current_password: str


class GenerateBlogDraftRequest(BaseModel):
    asin: str
    israel_price: float | None = None
    amazon_price: float
    manual_title: str | None = None
    manual_features: list[str] | None = None
    manual_brand: str | None = None
    manual_category: str | None = None
    manual_image: str | None = None
    min_order_49: bool = False
    voltage_warning: bool = False


router = APIRouter(prefix="/admin", tags=["admin"])


@router.get("/stats")
async def get_stats(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    today = datetime.utcnow() - timedelta(hours=24)
    # 3 User counts collapsed into one round-trip via conditional aggregation,
    # then run concurrently with the Product / NotificationLog counts.
    user_stmt = select(
        func.count().filter(and_(User.is_admin == False, User.is_verified == True)).label("total_users"),
        func.count().filter(User.is_admin == True).label("total_admins"),
        func.count().filter(and_(User.is_admin == False, User.is_verified == False)).label("unverified"),
    ).select_from(User)
    prod_stmt = select(func.count()).select_from(Product).where(Product.source == "user")
    notif_stmt = select(func.count()).select_from(NotificationLog).where(NotificationLog.sent_at >= today)

    user_row, total_products, notifs_today = await asyncio.gather(
        _one(user_stmt), _scalar(prod_stmt), _scalar(notif_stmt)
    )
    return {
        "total_users": user_row.total_users,
        "total_admins": user_row.total_admins,
        "total_products": total_products,
        "notifications_24h": notifs_today,
        "unverified_users": user_row.unverified,
    }


@router.get("/analytics")
async def get_analytics(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    now = datetime.utcnow()
    V = User.is_verified == True

    # All 9 User-table counts collapsed into a single round-trip via conditional
    # aggregation (count(*) FILTER (WHERE ...)). total_registered has no filter
    # since the query is already scoped to non-admin users.
    user_stmt = select(
        func.count().filter(V).label("total_verified"),
        func.count().filter(and_(V, User.last_login_at >= now - timedelta(days=7))).label("active_7d"),
        func.count().filter(and_(V, User.last_login_at >= now - timedelta(days=30))).label("active_30d"),
        func.count().filter(and_(V, User.vacation_mode == True)).label("vacation_count"),
        func.count().filter(and_(V, or_(User.last_login_at <= now - timedelta(days=14), User.last_login_at == None))).label("churn_14"),
        func.count().filter(and_(V, or_(User.last_login_at <= now - timedelta(days=30), User.last_login_at == None))).label("churn_30"),
        func.count().filter(and_(V, User.notify_email_bounced == True)).label("bounce_count"),
        func.count().filter(and_(V, User.google_id != None)).label("google_users"),
        func.count().label("total_registered"),
    ).select_from(User).where(User.is_admin == False)

    prod_dist_stmt = select(UserProduct.user_id, func.count().label("cnt")).group_by(UserProduct.user_id)
    send_log_stmt = (
        select(EmailSendLog.template_id, EmailSendLog.template_name,
               func.sum(EmailSendLog.sent_count).label("sent"))
        .group_by(EmailSendLog.template_id, EmailSendLog.template_name)
        .order_by(func.sum(EmailSendLog.sent_count).desc())
    )
    open_stmt = select(EmailOpen.template_id, func.count().label("opens")).group_by(EmailOpen.template_id)
    clicks_stmt = select(func.count()).select_from(EmailClick)

    # 5 independent round-trips run concurrently instead of 13 sequential ones.
    u, prod_dist_rows, send_log_rows, open_rows, total_clicks = await asyncio.gather(
        _one(user_stmt), _all(prod_dist_stmt), _all(send_log_stmt), _all(open_stmt), _scalar(clicks_stmt)
    )

    total_verified = u.total_verified
    active_7d = u.active_7d
    active_30d = u.active_30d
    vacation_count = u.vacation_count
    churn_risk_14d = u.churn_14
    churn_risk_30d = u.churn_30
    bounce_count = u.bounce_count
    google_users = u.google_users
    total_registered = u.total_registered

    dist_one = sum(1 for r in prod_dist_rows if r.cnt == 1)
    dist_two_five = sum(1 for r in prod_dist_rows if 2 <= r.cnt <= 5)
    dist_six_plus = sum(1 for r in prod_dist_rows if r.cnt >= 6)

    total_sent = sum((r.sent or 0) for r in send_log_rows)
    total_opens = sum(r.opens for r in open_rows)
    open_by_tid = {r.template_id: r.opens for r in open_rows}

    by_template = []
    for r in send_log_rows:
        sent = r.sent or 0
        opens = open_by_tid.get(r.template_id, 0)
        open_rate = round(opens / sent * 100, 1) if sent else 0
        by_template.append({
            "name": r.template_name,
            "sent": sent,
            "opens": opens,
            "open_rate": open_rate,
        })

    # Funnel — total_registered and google_users already computed above.
    return {
        "engagement": {
            "active_7d": active_7d,
            "active_30d": active_30d,
            "total_verified": total_verified,
            "vacation_count": vacation_count,
            "vacation_pct": round(vacation_count / total_verified * 100, 1) if total_verified else 0,
            "churn_risk_14d": churn_risk_14d,
            "churn_risk_30d": churn_risk_30d,
            "product_dist": {"one": dist_one, "two_to_five": dist_two_five, "six_plus": dist_six_plus},
        },
        "email": {
            "total_sent": total_sent,
            "total_opens": total_opens,
            "total_clicks": total_clicks,
            "open_rate": round(total_opens / total_sent * 100, 1) if total_sent else 0,
            "ctr": round(total_clicks / total_opens * 100, 1) if total_opens else 0,
            "bounce_count": bounce_count,
            "bounce_rate": round(bounce_count / total_verified * 100, 1) if total_verified else 0,
            "by_template": by_template,
        },
        "funnel": {
            "total_registered": total_registered,
            "verified": total_verified,
            "unverified": total_registered - total_verified,
            "verify_rate": round(total_verified / total_registered * 100, 1) if total_registered else 0,
            "google_users": google_users,
            "email_users": total_verified - google_users,
        },
    }


@router.get("/users")
async def list_users(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(User).where(User.is_verified == True, User.is_admin == False).order_by(User.created_at.desc()))
    users = result.scalars().all()

    # Batch-fetch product counts (one query instead of N)
    count_rows = await db.execute(
        select(UserProduct.user_id, func.count().label("cnt"))
        .group_by(UserProduct.user_id)
    )
    product_count_map = {row.user_id: row.cnt for row in count_rows}

    # Batch-fetch last product added date per user
    last_added_rows = await db.execute(
        select(UserProduct.user_id, func.max(UserProduct.added_at).label("last_added"))
        .group_by(UserProduct.user_id)
    )
    last_added_map = {row.user_id: row.last_added for row in last_added_rows}

    return [
        {
            "id": u.id,
            "email": u.email,
            "notify_email": u.notify_email,
            "is_active": u.is_active,
            "is_admin": u.is_admin,
            "is_verified": u.is_verified,
            "created_at": u.created_at.isoformat() if u.created_at else None,
            "product_count": product_count_map.get(u.id, 0),
            "max_products": u.max_products,
            "vacation_mode": u.vacation_mode,
            "notify_email_bounced": u.notify_email_bounced,
            "notify_email_bounce_type": u.notify_email_bounce_type,
            "notify_email_bounced_at": u.notify_email_bounced_at.isoformat() if u.notify_email_bounced_at else None,
            "last_login_at": u.last_login_at.isoformat() if u.last_login_at else None,
            "last_product_added_at": last_added_map[u.id].isoformat() if last_added_map.get(u.id) else None,
        }
        for u in users
    ]


@router.get("/users/{user_id}/products")
async def get_user_products(
    user_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(
        select(Product, UserProduct.custom_name, UserProduct.is_paused, UserProduct.added_at)
        .join(UserProduct, Product.id == UserProduct.product_id)
        .where(UserProduct.user_id == user_id)
        .order_by(UserProduct.added_at.desc())
    )
    return [
        {
            "asin": p.asin,
            "name": p.name or "",
            "url": p.url,
            "last_status": p.last_status,
            "last_checked": p.last_checked.isoformat() if p.last_checked else None,
            "custom_name": custom_name,
            "is_paused": is_paused,
            "added_at": added_at.isoformat() if added_at else None,
            "last_price": p.last_price or "",
        }
        for p, custom_name, is_paused, added_at in result.all()
    ]


@router.get("/users/{user_id}/email-history")
async def get_user_email_history(
    user_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    rows = (await db.execute(
        select(EmailSendRecipient, EmailSendLog)
        .join(EmailSendLog, EmailSendRecipient.send_log_id == EmailSendLog.id)
        .where(EmailSendRecipient.user_id == user_id)
        .order_by(EmailSendLog.sent_at.desc())
        .limit(30)
    )).all()

    if not rows:
        return []

    log_ids = [log.id for _, log in rows]
    clicked_result = await db.execute(
        select(EmailSendLog.id)
        .join(EmailSendRecipient, EmailSendRecipient.send_log_id == EmailSendLog.id)
        .join(EmailClick, (EmailClick.user_id == user_id) & (EmailClick.clicked_at >= EmailSendLog.sent_at))
        .where(EmailSendLog.id.in_(log_ids), EmailSendRecipient.user_id == user_id)
        .distinct()
    )
    clicked_log_ids = {r for r, in clicked_result.all()}

    return [
        {
            "template_name": log.template_name,
            "sent_at": (log.sent_at + timedelta(hours=3)).strftime("%d/%m/%Y %H:%M") if log.sent_at else "",
            "success": recipient.success,
            "clicked": log.id in clicked_log_ids,
        }
        for recipient, log in rows
    ]


@router.patch("/users/{user_id}/toggle-active")
async def toggle_active(
    user_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_active = not user.is_active
    await db.commit()
    return {"id": user.id, "is_active": user.is_active}


@router.patch("/users/{user_id}/toggle-admin")
async def toggle_admin(
    user_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_admin = not user.is_admin
    await db.commit()
    return {"id": user.id, "is_admin": user.is_admin}


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    if user_id == admin.id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    # Explicitly delete related records to avoid FK issues on older DB schemas
    await db.execute(delete(UserProduct).where(UserProduct.user_id == user_id))
    await db.execute(delete(NotificationLog).where(NotificationLog.user_id == user_id))
    await db.execute(delete(EmailClick).where(EmailClick.user_id == user_id))
    await db.execute(delete(EmailOpen).where(EmailOpen.user_id == user_id))
    await db.execute(delete(User).where(User.id == user_id))
    await db.commit()
    return {"deleted": user_id}


@router.get("/products")
async def list_products(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Product).where(Product.source == "user").order_by(Product.last_checked.desc().nullslast()))
    products = result.scalars().all()

    # Batch-fetch watcher counts (one query instead of N)
    watcher_rows = await db.execute(
        select(UserProduct.product_id, func.count().label("cnt"))
        .group_by(UserProduct.product_id)
    )
    watcher_map = {row.product_id: row.cnt for row in watcher_rows}

    # Batch-fetch paused watcher counts
    paused_rows = await db.execute(
        select(UserProduct.product_id, func.count().label("cnt"))
        .where(UserProduct.is_paused == True)
        .group_by(UserProduct.product_id)
    )
    paused_map = {row.product_id: row.cnt for row in paused_rows}

    # Count notifications in current streak: since GREATEST(free_since, last_click)
    # Matches the auto-pause countdown logic exactly
    notif_rows = await db.execute(text("""
        SELECT nl.product_id, COUNT(*) AS cnt
        FROM notification_log nl
        JOIN products p ON p.id = nl.product_id
        WHERE nl.success = TRUE
          AND nl.sent_at > GREATEST(
              COALESCE(p.free_since, '1970-01-01T00:00:00+00:00'::timestamptz),
              COALESCE(
                  (SELECT MAX(ec.clicked_at) FROM email_clicks ec
                   WHERE ec.user_id = nl.user_id AND ec.asin = p.asin),
                  '1970-01-01T00:00:00+00:00'::timestamptz
              )
          )
        GROUP BY nl.product_id
    """))
    notif_map = {row.product_id: row.cnt for row in notif_rows}

    # Total notification count ever per product
    notif_total_rows = await db.execute(
        select(NotificationLog.product_id, func.count().label("cnt"))
        .where(NotificationLog.success == True)
        .group_by(NotificationLog.product_id)
    )
    notif_total_map = {row.product_id: row.cnt for row in notif_total_rows}

    return [
        {
            "id": p.id,
            "asin": p.asin,
            "name": p.name,
            "url": p.url,
            "last_status": p.last_status,
            "last_checked": p.last_checked.isoformat() if p.last_checked else None,
            "consecutive_errors": p.consecutive_errors,
            "watchers": watcher_map.get(p.id, 0),
            "paused_watchers": paused_map.get(p.id, 0),
            "raw_text": p.raw_text[:200] if p.raw_text else "",
            "last_price": p.last_price or "",
            "image_url": p.image_url or f"https://images-na.ssl-images-amazon.com/images/P/{p.asin}.01._SL100_.jpg",
            "screenshot_path": os.path.basename(p.screenshot_path) if p.screenshot_path else None,
            "status_since": p.status_since.isoformat() if p.status_since else None,
            "notification_count": notif_map.get(p.id, 0),
            "notification_count_total": notif_total_map.get(p.id, 0),
        }
        for p in products
    ]


@router.get("/registrations-chart")
async def registrations_chart(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(
        select(cast(User.created_at, Date).label("date"), func.count().label("count"))
        .where(User.is_admin == False)
        .group_by(cast(User.created_at, Date))
        .order_by(cast(User.created_at, Date).asc())
        .limit(30)
    )
    return [{"date": str(row.date), "count": row.count} for row in result.all()]


@router.get("/notifications-log")
async def notifications_log(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
    limit: int = 50,
):
    result = await db.execute(
        select(NotificationLog, User.email, Product.name, Product.asin)
        .join(User, NotificationLog.user_id == User.id)
        .join(Product, NotificationLog.product_id == Product.id)
        .order_by(NotificationLog.sent_at.desc())
        .limit(limit)
    )
    return [
        {
            "sent_at": log.sent_at.isoformat(),
            "user_email": email,
            "email_to": log.email_to,
            "product_name": name or asin,
            "asin": asin,
            "status": log.status,
            "success": log.success,
            "error_msg": log.error_msg,
        }
        for log, email, name, asin in result.all()
    ]


@router.get("/system-message")
async def get_system_message(db: Annotated[AsyncSession, Depends(get_db)]):
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "system_message"))).scalar_one_or_none()
    return {"message": row.value if row else ""}


@router.post("/system-message")
async def set_system_message(
    body: dict,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    msg = str(body.get("message", "")).strip()
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "system_message"))).scalar_one_or_none()
    if row:
        row.value = msg
    else:
        db.add(SystemSetting(key="system_message", value=msg))
    await db.commit()
    return {"message": msg}


@router.get("/global-product-limit")
async def get_global_product_limit(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "max_products_per_user"))).scalar_one_or_none()
    return {"limit": int(row.value) if row else 10}


@router.post("/global-product-limit")
async def set_global_product_limit(
    body: dict,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    limit = int(body.get("limit", 20))
    if limit < 1 or limit > 10000:
        raise HTTPException(status_code=400, detail="מגבלה לא חוקית (1–10000)")
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "max_products_per_user"))).scalar_one_or_none()
    if row:
        row.value = str(limit)
    else:
        db.add(SystemSetting(key="max_products_per_user", value=str(limit)))
    await db.commit()
    return {"limit": limit, "message": f"מגבלת מוצרים גלובלית עודכנה ל-{limit}"}


@router.patch("/users/{user_id}/product-limit")
async def set_user_product_limit(
    user_id: int,
    body: dict,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    raw = body.get("limit")
    if raw is None or raw == "":
        user.max_products = None  # revert to global
    else:
        val = int(raw)
        if val < 1 or val > 10000:
            raise HTTPException(status_code=400, detail="מגבלה לא חוקית")
        user.max_products = val
    await db.commit()
    return {"user_id": user_id, "max_products": user.max_products}


@router.get("/inactivity-days")
async def get_inactivity_days(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "inactivity_days"))).scalar_one_or_none()
    return {"days": int(row.value) if row else 90}


@router.post("/inactivity-days")
async def set_inactivity_days(
    body: dict,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    days = int(body.get("days", 90))
    if days < 0 or days > 3650:
        raise HTTPException(status_code=400, detail="ערך לא חוקי (0–3650)")
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "inactivity_days"))).scalar_one_or_none()
    if row:
        row.value = str(days)
    else:
        db.add(SystemSetting(key="inactivity_days", value=str(days)))
    await db.commit()
    msg = f"מעבר למצב חופשה אחרי {days} ימי חוסר פעילות" if days > 0 else "בדיקת חוסר פעילות מושבתת"
    return {"days": days, "message": msg}


@router.post("/trigger-summary")
async def trigger_summary(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
    to: str | None = None,
    asins: str | None = None,
):
    """Send daily summary to a specific email (or admin if none).
    asins: optional comma-separated ASINs to override the product list (e.g. B08NYMBXG8,B07QXV6N1B).
    """
    from backend.models import Product, UserProduct
    from backend.notifier import send_daily_summary

    target = admin
    if to:
        found = (await db.execute(select(User).where(User.email == to))).scalar_one_or_none()
        if not found:
            found = (await db.execute(select(User).where(User.notify_email == to))).scalar_one_or_none()
        if found:
            target = found

    if asins:
        asin_list = [a.strip().upper() for a in asins.split(",") if a.strip()]
        products_result = await db.execute(
            select(Product).where(Product.asin.in_(asin_list))
        )
        db_products = {p.asin: p for p in products_result.scalars().all()}
        products = []
        for asin in asin_list:
            if asin in db_products:
                products.append((db_products[asin], None))
            else:
                dummy = Product()
                dummy.asin = asin
                dummy.name = f"מוצר לדוגמא — {asin}"
                dummy.url = f"https://www.amazon.com/dp/{asin}"
                dummy.last_status = "free"
                products.append((dummy, None))
    else:
        products_result = await db.execute(
            select(Product, UserProduct.custom_name)
            .join(UserProduct, Product.id == UserProduct.product_id)
            .where(UserProduct.user_id == target.id, UserProduct.is_paused == False)
        )
        products = products_result.all()

    if not products:
        dummy = Product()
        dummy.asin = "B08NYMBXG8"
        dummy.name = "מוצר לדוגמא"
        dummy.url = "https://www.amazon.com/dp/B08NYMBXG8"
        dummy.last_status = "free"
        products = [(dummy, None)]

    dest = to or target.notify_email
    target.notify_email = dest
    send_daily_summary(target, products)
    return {"message": f"✅ סיכום נשלח ל-{dest} ({len(products)} מוצרים)"}


@router.get("/get-check-time")
async def get_check_time(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "check_time"))).scalar_one_or_none()
    return {"time": row.value if row else "06:00"}


@router.post("/set-check-time")
async def set_check_time(
    body: dict,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    time_str = body.get("time", "")
    try:
        h, m = time_str.split(":")
        h, m = int(h), int(m)
        assert 0 <= h <= 23 and 0 <= m <= 59
    except Exception:
        raise HTTPException(status_code=400, detail="פורמט שגוי. נדרש HH:MM")
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "check_time"))).scalar_one_or_none()
    if row:
        row.value = time_str
    else:
        db.add(SystemSetting(key="check_time", value=time_str))
    await db.commit()
    from backend.main import reschedule_check_job
    reschedule_check_job(h, m)
    return {"time": time_str, "message": f"בדיקה יומית עודכנה ל-{time_str}"}


@router.post("/trigger-automation")
async def trigger_automation(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Trigger automation emails immediately (for testing)."""
    from backend.scheduler import run_automation_emails
    import asyncio
    asyncio.create_task(run_automation_emails())
    return {"message": "Automation emails triggered"}


@router.post("/users/{user_id}/clear-bounce")
async def clear_user_bounce(
    user_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.notify_email_bounced = False
    user.notify_email_bounced_at = None
    user.notify_email_bounce_type = None
    await db.commit()
    return {"message": f"Bounce cleared for user {user_id}"}


@router.post("/users/{user_id}/reset-automation")
async def reset_user_automation(
    user_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Reset automation flags for a user (for testing)."""
    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.automation_activation_sent_at = None
    user.automation_reminder_sent_at = None
    user.automation_expansion_sent_at = None
    await db.commit()
    return {"message": f"Automation flags reset for user {user_id}"}


@router.post("/trigger-check")
async def trigger_check(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from backend.scheduler import run_global_check_cycle
    import asyncio
    asyncio.create_task(run_global_check_cycle())
    return {"message": "Check cycle triggered"}


@router.post("/products/{product_id}/check")
async def trigger_single_product_check(
    product_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    import asyncio
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    from backend.scheduler import check_single_product
    asyncio.create_task(check_single_product(product.asin, product.url))
    return {"message": f"Check triggered for {product.asin}"}


@router.post("/clear-cookies")
async def clear_cookies(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Clear session cookies from memory and DB."""
    from backend.checker import browser_manager
    browser_manager._session_cookies = []
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "amazon_session_cookies"))).scalar_one_or_none()
    if row:
        await db.delete(row)
        await db.commit()
    return {"message": "Cookies cleared"}


@router.get("/cookie-status")
async def cookie_status(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Return current session cookie state loaded in the running checker."""
    from backend.checker import browser_manager
    count = len(browser_manager._session_cookies)
    ts_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == "amazon_session_cookies_updated_at")
    )).scalar_one_or_none()
    return {
        "loaded": count > 0,
        "count": count,
        "updated_at": ts_row.value if ts_row else None,
    }


@router.post("/inject-cookies")
async def inject_cookies(
    body: dict,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Inject Amazon session cookies into the running checker.
    Accepts: {"cookies": [{"name": "...", "value": "..."}]} (JSON array from browser export)
    """
    import json
    from backend.checker import browser_manager

    raw = body.get("cookies", [])
    if not raw:
        raise HTTPException(status_code=400, detail="cookies array required")

    # Support both [{name, value}] and {name: value} formats
    if isinstance(raw, list):
        cookie_list = [{"name": c["name"], "value": c["value"]} for c in raw if "name" in c and "value" in c]
    elif isinstance(raw, dict):
        cookie_list = [{"name": k, "value": v} for k, v in raw.items()]
    else:
        raise HTTPException(status_code=400, detail="Invalid cookies format")

    # Persist cookies + injection timestamp to DB so they survive restarts
    from datetime import datetime, timezone
    now_iso = datetime.now(timezone.utc).isoformat()

    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "amazon_session_cookies"))).scalar_one_or_none()
    if row:
        row.value = json.dumps(cookie_list)
    else:
        db.add(SystemSetting(key="amazon_session_cookies", value=json.dumps(cookie_list)))

    ts_row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "amazon_session_cookies_updated_at"))).scalar_one_or_none()
    if ts_row:
        ts_row.value = now_iso
    else:
        db.add(SystemSetting(key="amazon_session_cookies_updated_at", value=now_iso))
    await db.commit()

    browser_manager._session_cookies = cookie_list
    return {"injected": len(cookie_list), "message": f"הוזרקו {len(cookie_list)} cookies — נשמרו ב-DB"}


@router.post("/products/{product_id}/reset-errors")
async def reset_product_errors(
    product_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    product.consecutive_errors = 0
    await db.commit()
    return {"asin": product.asin, "consecutive_errors": 0}


class BulkDeleteRequest(BaseModel):
    product_ids: list[int]


@router.delete("/products/bulk")
async def bulk_delete_products(
    body: BulkDeleteRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    if not body.product_ids:
        raise HTTPException(status_code=400, detail="לא נבחרו מוצרים")
    await db.execute(delete(Product).where(Product.id.in_(body.product_ids)))
    await db.commit()
    return {"deleted": len(body.product_ids)}


@router.delete("/products/{product_id}")
async def delete_product(
    product_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    await db.execute(delete(Product).where(Product.id == product_id))
    await db.commit()
    return {"deleted": product_id}


@router.post("/sync-vacation-pauses")
async def sync_vacation_pauses(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Pause all products of users currently in vacation mode."""
    from sqlalchemy import update as sa_update
    vacation_ids = (await db.execute(
        select(User.id).where(User.vacation_mode == True, User.is_admin == False)
    )).scalars().all()
    result = await db.execute(
        sa_update(UserProduct).where(UserProduct.user_id.in_(vacation_ids)).values(is_paused=True)
    )
    await db.commit()
    return {"vacation_users": len(vacation_ids), "products_paused": result.rowcount}


@router.delete("/products-orphans")
async def delete_orphan_products(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Delete all user-source products with 0 watchers. Scanner products are never deleted here."""
    watched_ids = select(UserProduct.product_id).distinct()
    orphans_result = await db.execute(
        select(Product.id, Product.asin).where(
            Product.source == "user",
            Product.id.not_in(watched_ids),
        )
    )
    orphans = orphans_result.all()
    if not orphans:
        return {"deleted": [], "count": 0}
    orphan_ids = [row.id for row in orphans]
    await db.execute(delete(NotificationLog).where(NotificationLog.product_id.in_(orphan_ids)))
    await db.execute(delete(Product).where(Product.id.in_(orphan_ids)))
    await db.commit()
    return {"deleted": [row.asin for row in orphans], "count": len(orphans)}


# ── Admin profile management ──────────────────────────────────────────────────

@router.patch("/profile/password")
async def change_password(
    body: ChangePasswordRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    if not verify_password(body.current_password, admin.password_hash):
        raise HTTPException(status_code=400, detail="הסיסמה הנוכחית שגויה")
    if len(body.new_password) < 6:
        raise HTTPException(status_code=400, detail="הסיסמה החדשה קצרה מדי (מינימום 6 תווים)")
    admin.password_hash = hash_password(body.new_password)
    await db.commit()
    return {"message": "הסיסמה עודכנה בהצלחה"}


@router.post("/reclassify-all")
async def reclassify_all_products(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Re-classify all products using stored raw_text + last_price (no Amazon fetch)."""
    from backend.checker import _classify, ShippingStatus

    result = await db.execute(select(Product).where(Product.source == "user"))
    products = result.scalars().all()

    changed = 0
    details = []
    for p in products:
        old_status = p.last_status
        raw = p.raw_text or ""
        price = p.last_price or ""

        if raw:
            new_status = _classify(raw)
        else:
            new_status = ShippingStatus.PAID if price.upper().startswith("ILS") else ShippingStatus(old_status)

        # Apply ILS override
        if price.upper().startswith("ILS") and new_status not in (ShippingStatus.FREE,):
            new_status = ShippingStatus.PAID

        if new_status.value != old_status:
            details.append({"asin": p.asin, "from": old_status, "to": new_status.value})
            p.last_status = new_status.value
            changed += 1

    await db.commit()
    return {"total": len(products), "changed": changed, "details": details}


@router.post("/profile/request-email-change")
async def request_email_change(
    body: RequestEmailChangeRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from jose import jwt as jose_jwt
    from datetime import datetime, timedelta
    import os

    if not verify_password(body.current_password, admin.password_hash):
        raise HTTPException(status_code=400, detail="הסיסמה שגויה")

    # Check new email not already taken
    existing = (await db.execute(select(User).where(User.email == body.new_email))).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail="אימייל זה כבר בשימוש")

    # Create verification token (valid 1 hour)
    expire = datetime.utcnow() + timedelta(hours=1)
    token = jose_jwt.encode(
        {"sub": str(admin.id), "new_email": body.new_email, "type": "email_change", "exp": expire},
        SECRET_KEY, algorithm=ALGORITHM
    )

    base_url = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "app.amzfreeil.com")
    verify_url = f"https://{base_url}/admin/verify-email?token={token}"

    from backend.notifier import send_simple_email
    html = f"""
    <div dir="rtl" style="font-family:Arial,sans-serif; max-width:480px; margin:auto; padding:24px;">
      <h2 style="color:#e47911;">אימות שינוי אימייל · Amazon Israel Alert</h2>
      <p>קיבלנו בקשה לשנות את כתובת האימייל של חשבון המנהל שלך ל:</p>
      <p style="font-size:1.1rem; font-weight:bold; direction:ltr;">{body.new_email}</p>
      <p>לאישור השינוי לחץ על הכפתור:</p>
      <a href="{verify_url}" style="display:inline-block; background:#FF9900; color:#111;
         padding:12px 28px; border-radius:8px; font-weight:bold; text-decoration:none; margin:16px 0;">
        אשר שינוי אימייל
      </a>
      <p style="color:#888; font-size:0.85rem;">הקישור תקף לשעה אחת. אם לא ביקשת שינוי זה, התעלם.</p>
    </div>
    """
    send_simple_email(body.new_email, "אימות שינוי אימייל · Amazon Israel Alert", html)
    return {"message": f"קישור אימות נשלח ל-{body.new_email}"}


@router.get("/verify-email")
async def verify_email_change(
    token: str,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from jose import jwt as jose_jwt, JWTError
    try:
        payload = jose_jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("type") != "email_change":
            raise HTTPException(status_code=400, detail="טוקן לא תקין")
        user_id = int(payload["sub"])
        new_email = payload["new_email"]
    except JWTError:
        raise HTTPException(status_code=400, detail="הקישור פג תוקף או לא תקין")

    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="משתמש לא נמצא")

    user.email = new_email
    user.notify_email = new_email
    await db.commit()
    from fastapi.responses import HTMLResponse
    return HTMLResponse("""
    <html dir="rtl"><body style="font-family:Arial; text-align:center; padding:60px; background:#fffaf1;">
    <h2 style="color:#2e7d32;">✅ האימייל עודכן בהצלחה!</h2>
    <p>כתובת האימייל שלך עודכנה. <a href="/admin/login">לחץ כאן לכניסה מחדש</a></p>
    </body></html>
    """)


@router.get("/checks-status")
async def checks_status(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "system_paused"))).scalar_one_or_none()
    paused = row is not None and row.value == "true"
    return {"paused": paused}


@router.post("/pause-checks")
async def pause_checks(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from backend.main import scheduler
    for job_id in ("global_check", "daily_summary"):
        job = scheduler.get_job(job_id)
        if job:
            scheduler.pause_job(job_id)
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "system_paused"))).scalar_one_or_none()
    if row:
        row.value = "true"
    else:
        db.add(SystemSetting(key="system_paused", value="true"))
    await db.commit()
    return {"paused": True, "message": "הבדיקות הושהו"}


@router.post("/resume-checks")
async def resume_checks(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from backend.main import scheduler
    for job_id in ("global_check", "daily_summary"):
        job = scheduler.get_job(job_id)
        if job:
            scheduler.resume_job(job_id)
    row = (await db.execute(select(SystemSetting).where(SystemSetting.key == "system_paused"))).scalar_one_or_none()
    if row:
        row.value = "false"
    else:
        db.add(SystemSetting(key="system_paused", value="false"))
    await db.commit()
    return {"paused": False, "message": "הבדיקות הופעלו מחדש"}


@router.post("/test-cookies")
async def test_cookies(
    body: dict,
    admin: Annotated[User, Depends(get_current_admin)],
):
    """Test if provided cookie string returns Israel location on Amazon.
    Accepts: {"cookies": "session-id=xxx; ubid-main=yyy; ..."}
    Returns: nav text and whether Israel was detected.
    """
    import httpx
    from bs4 import BeautifulSoup

    cookie_str = body.get("cookies", "").strip()
    if not cookie_str:
        raise HTTPException(status_code=400, detail="cookies field required")

    # Parse "name=value; name2=value2" format
    cookie_dict = {}
    for part in cookie_str.split(";"):
        part = part.strip()
        if "=" in part:
            k, _, v = part.partition("=")
            cookie_dict[k.strip()] = v.strip()

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        async with httpx.AsyncClient(
            headers=headers, cookies=cookie_dict,
            follow_redirects=True, timeout=20.0,
        ) as client:
            resp = await client.get("https://www.amazon.com/dp/B00EDR1X3O?psc=1&th=1")
            soup = BeautifulSoup(resp.text, "html.parser")
            nav_text = ""
            for nav_id in ["glow-ingress-line2", "glow-ingress-line1", "nav-global-location-popover-link"]:
                el = soup.find(id=nav_id)
                if el:
                    nav_text = el.get_text(strip=True)
                    break
            israel_detected = "israel" in nav_text.lower() or "israel" in resp.text.lower()
            return {
                "nav_text": nav_text,
                "israel_detected": israel_detected,
                "cookies_parsed": len(cookie_dict),
                "status_code": resp.status_code,
            }
    except Exception as e:
        return {"error": str(e), "israel_detected": False}


@router.get("/clicks")
async def get_click_analytics(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
    days: int = 7,
):
    since = datetime.utcnow() - timedelta(days=days)

    total = (
        await db.execute(
            select(func.count()).select_from(EmailClick).where(EmailClick.clicked_at >= since)
        )
    ).scalar()

    by_asin_rows = (
        await db.execute(
            select(EmailClick.asin, func.count().label("cnt"))
            .where(EmailClick.clicked_at >= since)
            .group_by(EmailClick.asin)
            .order_by(func.count().desc())
            .limit(20)
        )
    ).all()

    by_day_rows = (
        await db.execute(
            select(cast(EmailClick.clicked_at, Date).label("day"), func.count().label("cnt"))
            .where(EmailClick.clicked_at >= since)
            .group_by(cast(EmailClick.clicked_at, Date))
            .order_by(cast(EmailClick.clicked_at, Date))
        )
    ).all()

    recent_rows = (
        await db.execute(
            select(EmailClick, User.email)
            .outerjoin(User, EmailClick.user_id == User.id)
            .where(EmailClick.clicked_at >= since)
            .order_by(EmailClick.clicked_at.desc())
            .limit(50)
        )
    ).all()

    return {
        "total": total,
        "days": days,
        "by_asin": [{"asin": r.asin, "count": r.cnt} for r in by_asin_rows],
        "by_day": [{"date": str(r.day), "count": r.cnt} for r in by_day_rows],
        "recent": [
            {
                "id": r.EmailClick.id,
                "user_email": r.email or f"user#{r.EmailClick.user_id}",
                "asin": r.EmailClick.asin,
                "clicked_at": (r.EmailClick.clicked_at + timedelta(hours=3)).strftime("%d/%m/%Y %H:%M") if r.EmailClick.clicked_at else "",
                "ip": r.EmailClick.ip or "—",
            }
            for r in recent_rows
        ],
    }


@router.delete("/clicks/{click_id}")
async def delete_click(
    click_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    await db.execute(delete(EmailClick).where(EmailClick.id == click_id))
    await db.commit()
    return {"ok": True}


@router.post("/send-test-click-email")
async def send_test_click_email(
    admin: Annotated[User, Depends(get_current_admin)],
    to: str | None = None,
):
    import os
    from urllib.parse import urlencode
    from backend.notifier import _send_via_resend

    test_asin = "B0BG52SJ5N"
    base = os.environ.get("APP_BASE_URL", "https://app.amzfreeil.com").rstrip("/")
    affiliate_tag = os.environ.get("AMAZON_AFFILIATE_TAG", "").strip()
    dest = f"https://www.amazon.com/dp/{test_asin}?tag={affiliate_tag}" if affiliate_tag else f"https://www.amazon.com/dp/{test_asin}"
    params = urlencode({"u": admin.id, "a": test_asin, "url": dest})
    tracking = f"{base}/track/click?{params}"

    html = f"""
    <div dir="rtl" style="font-family:Arial,sans-serif;max-width:480px;margin:auto;padding:24px;background:#fffaf1;border-radius:12px;">
      <h2 style="color:#e47911;">🧪 מייל בדיקה — Click Tracking</h2>
      <p style="color:#555;">לחץ על הכפתור ובדוק שמופיע click ב-<strong>/admin/clicks</strong></p>
      <a href="{tracking}"
         style="display:inline-block;background:#FF9900;color:#111;padding:12px 28px;border-radius:8px;font-weight:bold;text-decoration:none;margin-top:16px;">
        קנה עכשיו — משלוח חינם (בדיקה)
      </a>
      <p style="margin-top:16px;font-size:12px;color:#999;">ASIN: {test_asin} · user_id: {admin.id}</p>
    </div>"""

    dest = to or admin.notify_email
    ok = _send_via_resend(dest, "🧪 בדיקת Click Tracking — amzfreeil", html, f"לחץ כאן: {tracking}")
    return {"sent": ok, "to": dest, "tracking_url": tracking}


@router.post("/fix-google-proxy-opens")
async def fix_google_proxy_opens(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """One-time migration: mark all Google IP opens as suspicious."""
    from sqlalchemy import update as sa_update
    result = await db.execute(
        sa_update(EmailOpen)
        .where(EmailOpen.ip.like("66.249.%"), EmailOpen.is_suspicious == False)
        .values(is_suspicious=True)
    )
    await db.commit()
    return {"updated": result.rowcount}


@router.post("/send-test-newsletter")
async def send_test_newsletter(
    admin: Annotated[User, Depends(get_current_admin)],
    to: str | None = None,
):
    from backend.notifier import send_newsletter_test, _pause_url
    target = to or admin.notify_email or admin.email
    pause = _pause_url(admin.id)
    ok = send_newsletter_test(target, pause)
    return {"ok": ok, "to": target}


@router.post("/seed-newsletter-template")
async def seed_newsletter_template(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from backend.notifier import build_newsletter_html, _NEWSLETTER_SUBJECT
    name = "עדכון_מוצר_מאי_2026"
    body = build_newsletter_html("{{pause_url}}")
    existing = (await db.execute(select(EmailTemplate).where(EmailTemplate.name == name))).scalar_one_or_none()
    if existing:
        existing.subject = _NEWSLETTER_SUBJECT
        existing.body = body
        await db.commit()
        return {"id": existing.id, "message": "תבנית עודכנה", "already_exists": True}
    t = EmailTemplate(name=name, subject=_NEWSLETTER_SUBJECT, body=body)
    db.add(t)
    await db.commit()
    await db.refresh(t)
    return {"id": t.id, "message": "תבנית ניוזלטר נוצרה", "already_exists": False}


@router.post("/send-test-telegram-invite")
async def send_test_telegram_invite(
    admin: Annotated[User, Depends(get_current_admin)],
    to: str | None = None,
):
    from backend.notifier import send_telegram_invite_test, _pause_url
    target = to or admin.notify_email or admin.email
    pause = _pause_url(admin.id)
    ok = send_telegram_invite_test(target, pause)
    return {"ok": ok, "to": target}


@router.post("/seed-telegram-invite-template")
async def seed_telegram_invite_template(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from backend.notifier import build_telegram_invite_html, _TELEGRAM_INVITE_SUBJECT
    name = "הזמנה_טלגרם_יוני_2026"
    body = build_telegram_invite_html("{{pause_url}}")
    existing = (await db.execute(select(EmailTemplate).where(EmailTemplate.name == name))).scalar_one_or_none()
    if existing:
        existing.subject = _TELEGRAM_INVITE_SUBJECT
        existing.body = body
        await db.commit()
        return {"id": existing.id, "message": "תבנית עודכנה", "already_exists": True}
    t = EmailTemplate(name=name, subject=_TELEGRAM_INVITE_SUBJECT, body=body)
    db.add(t)
    await db.commit()
    await db.refresh(t)
    return {"id": t.id, "message": "תבנית הזמנת טלגרם נוצרה", "already_exists": False}


@router.get("/quick-log")
async def quick_log(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
    limit: int = 50,
):
    """Fast unified log: last N email send events with per-recipient details."""
    from backend.models import EmailSendLog, EmailSendRecipient
    from sqlalchemy import select
    from datetime import timezone, timedelta
    israel = timezone(timedelta(hours=3))

    logs = (await db.execute(
        select(EmailSendLog).order_by(EmailSendLog.sent_at.desc()).limit(limit)
    )).scalars().all()

    log_ids = [l.id for l in logs]
    recipients = (await db.execute(
        select(EmailSendRecipient).where(EmailSendRecipient.send_log_id.in_(log_ids))
    )).scalars().all()

    recip_by_log: dict[int, list] = {}
    for r in recipients:
        recip_by_log.setdefault(r.send_log_id, []).append({
            "email": r.email,
            "success": r.success,
        })

    return [
        {
            "id": l.id,
            "template": l.template_name,
            "audience": l.audience,
            "sent_at": l.sent_at.astimezone(israel).strftime("%d/%m/%Y %H:%M") if l.sent_at else None,
            "sent": l.sent_count,
            "failed": l.failed_count,
            "recipients": recip_by_log.get(l.id, []),
        }
        for l in logs
    ]


# ── Email Templates ───────────────────────────────────────────────────────────

class EmailTemplateBody(BaseModel):
    name: str
    subject: str
    body: str

class EmailTemplateSendBody(BaseModel):
    audience: str  # "all" | "active" | "vacation" | "inactive" | "single" | "custom"
    user_id: int | None = None
    products_min: int | None = None  # include users with >= this many products
    products_max: int | None = None  # include users with <= this many products
    custom_emails: list[str] | None = None  # for audience="custom"
    to_email: str | None = None  # override recipient for audience="self"


@router.get("/email-templates")
async def list_email_templates(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(EmailTemplate).order_by(EmailTemplate.created_at.desc()))
    templates = result.scalars().all()
    return [
        {"id": t.id, "name": t.name, "subject": t.subject, "body": t.body,
         "created_at": t.created_at.isoformat() if t.created_at else None}
        for t in templates
    ]


@router.post("/email-templates")
async def create_email_template(
    body: EmailTemplateBody,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    existing = (await db.execute(select(EmailTemplate).where(EmailTemplate.name == body.name))).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail="שם תבנית כבר קיים")
    t = EmailTemplate(name=body.name.strip(), subject=body.subject.strip(), body=body.body)
    db.add(t)
    await db.commit()
    await db.refresh(t)
    return {"id": t.id, "message": "תבנית נשמרה"}


@router.put("/email-templates/{template_id}")
async def update_email_template(
    template_id: int,
    body: EmailTemplateBody,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    t = (await db.execute(select(EmailTemplate).where(EmailTemplate.id == template_id))).scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לא נמצאה")
    # Check name uniqueness if changed
    if body.name.strip() != t.name:
        dup = (await db.execute(select(EmailTemplate).where(EmailTemplate.name == body.name.strip()))).scalar_one_or_none()
        if dup:
            raise HTTPException(status_code=400, detail="שם תבנית כבר קיים")
    t.name = body.name.strip()
    t.subject = body.subject.strip()
    t.body = body.body
    await db.commit()
    return {"message": "תבנית עודכנה"}


@router.delete("/email-templates/{template_id}")
async def delete_email_template(
    template_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    t = (await db.execute(select(EmailTemplate).where(EmailTemplate.id == template_id))).scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לא נמצאה")
    await db.delete(t)
    await db.commit()
    return {"message": "תבנית נמחקה"}


# In-memory job tracker for send progress
_send_jobs: dict[str, dict] = {}


async def _execute_send_job(
    job_id: str,
    template_id: int,
    tpl_name: str,
    tpl_subject: str,
    tpl_body: str,
    audience: str,
    base_url: str,
    user_data: list,  # list of (user_id, email, notify_email, pc)
):
    import asyncio
    from backend.notifier import _send_via_resend, _pause_url
    from backend.database import AsyncSessionLocal

    job = _send_jobs[job_id]
    sent = failed = 0
    recipients_to_save = []

    from urllib.parse import quote as _quote

    for i, (uid, email, notify_email, pc) in enumerate(user_data):
        recipient = notify_email or email
        subj = tpl_subject.replace("{{email}}", email).replace("{{product_count}}", str(pc))

        def _turl(dest: str) -> str:
            return f"{base_url}/track/click?u={uid}&a=&url={_quote(dest, safe='')}"

        open_pixel = (
            f'<img src="{base_url}/track/email-open?uid={uid}&tn={tpl_name}"'
            ' width="1" height="1" style="display:none;border:0;" alt="">'
        )
        html_body = (
            tpl_body
            .replace("{{email}}", email)
            .replace("{{notify_email}}", recipient)
            .replace("{{product_count}}", str(pc))
            .replace("{{pause_url}}", _pause_url(uid))
            .replace("{{open_pixel}}", open_pixel)
            .replace("{{track_search}}", _turl("https://www.amzfreeil.com/search.html"))
            .replace("{{track_free_products}}", _turl("https://www.amzfreeil.com/free-products.html"))
            .replace("{{track_dashboard}}", _turl("https://app.amzfreeil.com/dashboard"))
        )
        ok = _send_via_resend(recipient, subj, html_body, "")
        if ok:
            sent += 1
        else:
            failed += 1
        recipients_to_save.append((uid, recipient, ok))

        job["sent"] = sent
        job["failed"] = failed
        job["remaining"] = len(user_data) - (i + 1)

        await asyncio.sleep(0.55)

    # Persist to DB
    async with AsyncSessionLocal() as db:
        log = EmailSendLog(
            template_id=template_id,
            template_name=tpl_name,
            audience=audience,
            sent_count=sent,
            failed_count=failed,
        )
        db.add(log)
        await db.flush()
        for uid, email, ok in recipients_to_save:
            db.add(EmailSendRecipient(send_log_id=log.id, user_id=uid, email=email, success=ok))
        await db.commit()

    job["done"] = True
    job["message"] = f"נשלח ל-{sent} משתמשים" + (f", {failed} נכשלו" if failed else "")


@router.post("/email-templates/{template_id}/send")
async def send_email_template(
    template_id: int,
    body: EmailTemplateSendBody,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    import asyncio, uuid, os

    t = (await db.execute(select(EmailTemplate).where(EmailTemplate.id == template_id))).scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לא נמצאה")

    base_url = os.environ.get("APP_BASE_URL", "https://app.amzfreeil.com").rstrip("/")

    product_count_sub = (
        select(func.count(UserProduct.id))
        .where(UserProduct.user_id == User.id)
        .correlate(User)
        .scalar_subquery()
    )
    q = select(User, product_count_sub.label("pc")).where(User.is_verified == True, User.is_admin == False)

    if body.audience == "self":
        q = select(User, product_count_sub.label("pc")).where(User.id == admin.id)
    elif body.audience == "active":
        q = q.where(User.is_active == True, User.vacation_mode == False)
    elif body.audience == "vacation":
        q = q.where(User.is_active == True, User.vacation_mode == True)
    elif body.audience == "inactive":
        q = q.where(User.is_active == False)
    elif body.audience == "single":
        if not body.user_id:
            raise HTTPException(status_code=400, detail="חסר user_id")
        q = select(User, product_count_sub.label("pc")).where(User.id == body.user_id)
    elif body.audience == "custom":
        if not body.custom_emails:
            raise HTTPException(status_code=400, detail="חסרה רשימת מיילים")
        emails_clean = [e.strip().lower() for e in body.custom_emails if e.strip()]
        q = select(User, product_count_sub.label("pc")).where(
            User.is_verified == True,
            func.lower(User.notify_email).in_(emails_clean) | func.lower(User.email).in_(emails_clean)
        )

    if body.products_min is not None:
        q = q.where(product_count_sub >= body.products_min)
    if body.products_max is not None:
        q = q.where(product_count_sub <= body.products_max)

    rows = (await db.execute(q)).all()
    if not rows:
        return {"job_id": None, "total": 0, "message": "לא נמצאו משתמשים התואמים את הסינון"}

    # Extract plain data before session closes
    user_data = [(r[0].id, r[0].email, r[0].notify_email or r[0].email, r[1]) for r in rows]
    if body.audience == "self" and body.to_email:
        user_data = [(uid, email, body.to_email, pc) for uid, email, _, pc in user_data]

    job_id = str(uuid.uuid4())
    _send_jobs[job_id] = {
        "total": len(user_data), "sent": 0, "failed": 0,
        "remaining": len(user_data), "done": False, "message": "",
    }
    asyncio.create_task(_execute_send_job(
        job_id, template_id, t.name, t.subject, t.body, body.audience, base_url, user_data
    ))
    return {"job_id": job_id, "total": len(user_data)}


@router.get("/send-progress/{job_id}")
async def get_send_progress(
    job_id: str,
    admin: Annotated[User, Depends(get_current_admin)],
):
    job = _send_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@router.get("/email-send-logs")
async def list_send_logs(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
    days: int = 30,
    email: str | None = None,
):
    from datetime import datetime, timedelta, timezone
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)

    base_filter = [EmailSendLog.sent_at >= cutoff]
    if email:
        email_log_ids = select(EmailSendRecipient.send_log_id).where(
            EmailSendRecipient.email == email.strip().lower()
        )
        base_filter.append(EmailSendLog.id.in_(email_log_ids))

    # Main query: logs + clicks (unchanged logic)
    rows = (await db.execute(
        select(
            EmailSendLog,
            func.count(func.distinct(EmailClick.user_id)).label("clicks"),
        )
        .where(*base_filter)
        .outerjoin(EmailClick, (EmailClick.user_id.in_(
                                   select(EmailSendRecipient.user_id)
                                   .where(EmailSendRecipient.send_log_id == EmailSendLog.id,
                                          EmailSendRecipient.success == True)
                               )) &
                               (EmailClick.clicked_at >= EmailSendLog.sent_at))
        .group_by(EmailSendLog.id)
        .order_by(EmailSendLog.sent_at.desc())
        .limit(500)
    )).all()

    if not rows:
        return []

    # Separate query: opens per log (match by template_name + time window)
    log_ids = [r.EmailSendLog.id for r in rows]
    open_rows = (await db.execute(
        select(EmailSendRecipient.send_log_id, func.count(func.distinct(EmailOpen.user_id)).label("opens"))
        .join(EmailOpen, (EmailOpen.user_id == EmailSendRecipient.user_id) &
                         (EmailOpen.template_name == (
                             select(EmailSendLog.template_name)
                             .where(EmailSendLog.id == EmailSendRecipient.send_log_id)
                             .scalar_subquery()
                         )) &
                         (EmailOpen.opened_at >= (
                             select(EmailSendLog.sent_at)
                             .where(EmailSendLog.id == EmailSendRecipient.send_log_id)
                             .scalar_subquery()
                         )))
        .where(EmailSendRecipient.send_log_id.in_(log_ids), EmailSendRecipient.success == True)
        .group_by(EmailSendRecipient.send_log_id)
    )).all()
    opens_by_log = {r.send_log_id: r.opens for r in open_rows}

    return [
        {
            "id": r.EmailSendLog.id,
            "template_id": r.EmailSendLog.template_id,
            "template_name": r.EmailSendLog.template_name,
            "sent_at": r.EmailSendLog.sent_at.isoformat(),
            "audience": r.EmailSendLog.audience,
            "sent_count": r.EmailSendLog.sent_count,
            "failed_count": r.EmailSendLog.failed_count,
            "clicks": r.clicks,
            "opens": opens_by_log.get(r.EmailSendLog.id, 0),
        }
        for r in rows
    ]


@router.delete("/email-send-logs/{log_id}")
async def delete_send_log(
    log_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    await db.execute(delete(EmailSendRecipient).where(EmailSendRecipient.send_log_id == log_id))
    await db.execute(delete(EmailSendLog).where(EmailSendLog.id == log_id))
    await db.commit()
    return {"deleted": log_id}


@router.get("/email-send-logs/{log_id}/recipients")
async def get_send_log_recipients(
    log_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    log = (await db.execute(select(EmailSendLog).where(EmailSendLog.id == log_id))).scalar_one_or_none()
    if not log:
        raise HTTPException(status_code=404, detail="לוג לא נמצא")

    rows = (await db.execute(
        select(EmailSendRecipient)
        .where(EmailSendRecipient.send_log_id == log_id)
        .order_by(EmailSendRecipient.success.desc(), EmailSendRecipient.id)
    )).scalars().all()

    success_user_ids = [r.user_id for r in rows if r.user_id and r.success]
    clicked_ids: set[int] = set()
    opened_ids: set[int] = set()
    if success_user_ids:
        clicked_result = await db.execute(
            select(EmailClick.user_id)
            .where(
                EmailClick.user_id.in_(success_user_ids),
                EmailClick.clicked_at >= log.sent_at,
            )
            .distinct()
        )
        clicked_ids = {row[0] for row in clicked_result.all()}

        opened_result = await db.execute(
            select(EmailOpen.user_id, EmailOpen.is_suspicious)
            .where(
                EmailOpen.user_id.in_(success_user_ids),
                EmailOpen.template_name == log.template_name,
                EmailOpen.opened_at >= log.sent_at,
            )
        )
        # Per user: suspicious=False wins if any confirmed open exists
        opened_map: dict[int, bool] = {}
        for uid, is_susp in opened_result.all():
            if uid not in opened_map:
                opened_map[uid] = is_susp
            elif not is_susp:
                opened_map[uid] = False

    return [
        {
            "id": r.id, "user_id": r.user_id, "email": r.email,
            "success": r.success,
            "opened": (r.user_id in opened_map) if r.success else False,
            "opened_suspicious": opened_map.get(r.user_id, False) if r.success else False,
            "clicked": (r.user_id in clicked_ids) if r.success else False,
        }
        for r in rows
    ]


@router.post("/email-send-logs/{log_id}/resend-failed")
async def resend_failed_recipients(
    log_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from backend.notifier import _send_via_resend, _pause_url

    log = (await db.execute(select(EmailSendLog).where(EmailSendLog.id == log_id))).scalar_one_or_none()
    if not log:
        raise HTTPException(status_code=404, detail="לוג לא נמצא")
    if not log.template_id:
        raise HTTPException(status_code=400, detail="התבנית נמחקה — לא ניתן לשלוח מחדש")

    t = (await db.execute(select(EmailTemplate).where(EmailTemplate.id == log.template_id))).scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="תבנית לא נמצאה")

    failed_rows = (await db.execute(
        select(EmailSendRecipient)
        .where(EmailSendRecipient.send_log_id == log_id, EmailSendRecipient.success == False)
    )).scalars().all()

    if not failed_rows:
        return {"sent": 0, "failed": 0, "message": "אין נכשלים לשליחה מחדש"}

    import os, asyncio
    base_url = os.environ.get("APP_BASE_URL", "https://app.amzfreeil.com").rstrip("/")

    product_count_sub = (
        select(func.count(UserProduct.id))
        .where(UserProduct.user_id == User.id)
        .correlate(User)
        .scalar_subquery()
    )

    sent = failed = 0
    for i, r in enumerate(failed_rows):
        u = (await db.execute(select(User, product_count_sub.label("pc")).where(User.id == r.user_id))).first() if r.user_id else None
        recipient = r.email
        pc = u[1] if u else 0
        user_obj = u[0] if u else None
        subj = t.subject.replace("{{email}}", recipient).replace("{{product_count}}", str(pc))
        pixel_url = f"{base_url}/track/email-open?uid={r.user_id or 0}&tid={t.id}"
        pixel = f'<img src="{pixel_url}" width="1" height="1" style="display:none;" alt="">'
        pause = _pause_url(r.user_id) if r.user_id else "#"
        html_body = (
            t.body
            .replace("{{email}}", recipient)
            .replace("{{notify_email}}", recipient)
            .replace("{{product_count}}", str(pc))
            .replace("{{pause_url}}", pause)
        ) + pixel
        ok = _send_via_resend(recipient, subj, html_body, "")
        r.success = ok
        if ok:
            sent += 1
        else:
            failed += 1
        await asyncio.sleep(0.55)

    # Update log counts
    log.sent_count += sent
    log.failed_count -= sent  # sent successfully this time
    await db.commit()
    return {"sent": sent, "failed": failed, "message": f"נשלח מחדש ל-{sent} משתמשים" + (f", {failed} נכשלו שוב" if failed else "")}


@router.get("/email-templates/{template_id}/opens")
async def get_template_opens(
    template_id: int,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    rows = (await db.execute(
        select(EmailOpen, User.email)
        .join(User, User.id == EmailOpen.user_id)
        .where(EmailOpen.template_id == template_id)
        .order_by(EmailOpen.opened_at.desc())
    )).all()
    total_unique = len({r[0].user_id for r in rows})
    return {
        "total_opens": len(rows),
        "unique_openers": total_unique,
        "opens": [
            {"email": r[1], "opened_at": r[0].opened_at.isoformat(), "ip": r[0].ip}
            for r in rows
        ]
    }


# ─── Excel Export ─────────────────────────────────────────────────────────────

def _xl_header_style():
    fill = PatternFill(fill_type="solid", fgColor="E47911")
    font = Font(color="FFFFFF", bold=True)
    return fill, font

def _xl_title_style():
    return Font(bold=True, size=13)

def _xl_write_headers(ws, headers):
    fill, font = _xl_header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

def _xl_autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


@router.get("/export/excel")
async def export_excel(
    db: Annotated[AsyncSession, Depends(get_db)],
    token: str = Query(...),
):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

    result = await db.execute(
        select(User).where(User.id == int(user_id), User.is_active == True, User.is_admin == True)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Admin access required")

    now = datetime.utcnow()
    report_date = now.strftime("%d/%m/%Y %H:%M")
    file_date = now.strftime("%Y-%m-%d")

    wb = Workbook()

    # ── Sheet 1: סיכום ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "סיכום"
    ws.sheet_view.rightToLeft = True

    ws.append(["דוח מנהל — AmzFree Israel"])
    ws["A1"].font = _xl_title_style()
    ws.append([f"תאריך הפקה: {report_date}"])
    ws.append([])

    # Engagement data
    base_u = lambda *extra: select(func.count()).select_from(User).where(
        User.is_admin == False, User.is_verified == True, *extra
    )
    total_verified = (await db.execute(base_u())).scalar()
    active_7d  = (await db.execute(base_u(User.last_login_at >= now - timedelta(days=7)))).scalar()
    active_30d = (await db.execute(base_u(User.last_login_at >= now - timedelta(days=30)))).scalar()
    vacation   = (await db.execute(base_u(User.vacation_mode == True))).scalar()
    churn_14   = (await db.execute(base_u((User.last_login_at <= now - timedelta(days=14)) | (User.last_login_at == None)))).scalar()
    churn_30   = (await db.execute(base_u((User.last_login_at <= now - timedelta(days=30)) | (User.last_login_at == None)))).scalar()
    bounce_count = (await db.execute(base_u(User.notify_email_bounced == True))).scalar()
    google_users = (await db.execute(base_u(User.google_id != None))).scalar()
    total_reg  = (await db.execute(select(func.count()).select_from(User).where(User.is_admin == False))).scalar()
    unverified = total_reg - total_verified

    send_rows = (await db.execute(select(func.sum(EmailSendLog.sent_count)))).scalar() or 0
    total_opens_count = (await db.execute(select(func.count()).select_from(EmailOpen))).scalar()
    total_clicks_count = (await db.execute(select(func.count()).select_from(EmailClick))).scalar()
    open_rate = round(total_opens_count / send_rows * 100, 1) if send_rows else 0
    ctr       = round(total_clicks_count / total_opens_count * 100, 1) if total_opens_count else 0
    bounce_rate = round(bounce_count / total_verified * 100, 1) if total_verified else 0
    verify_rate = round(total_verified / total_reg * 100, 1) if total_reg else 0

    sections = [
        ("📊 מעורבות משתמשים", [
            ("סה\"כ משתמשים מאומתים", total_verified),
            ("פעילים — 7 ימים אחרונים", active_7d),
            ("פעילים — 30 ימים אחרונים", active_30d),
            ("מצב חופשה", vacation),
            ("בסכנת נטישה (14 יום)", churn_14),
            ("בסכנת נטישה (30 יום)", churn_30),
        ]),
        ("📧 ביצועי מיילים", [
            ("סה\"כ מיילים שנשלחו", send_rows),
            ("סה\"כ פתיחות", total_opens_count),
            ("Open Rate %", f"{open_rate}%"),
            ("סה\"כ לחיצות", total_clicks_count),
            ("CTR %", f"{ctr}%"),
            ("Bounces", bounce_count),
            ("Bounce Rate %", f"{bounce_rate}%"),
        ]),
        ("🔍 Funnel גיוס", [
            ("סה\"כ נרשמו", total_reg),
            ("אימתו אימייל", total_verified),
            ("לא אימתו", unverified),
            ("אחוז אימות", f"{verify_rate}%"),
            ("הרשמה דרך Google", google_users),
            ("הרשמה דרך אימייל+סיסמה", total_verified - google_users),
        ]),
    ]

    for section_title, rows_data in sections:
        ws.append([section_title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
        ws.append(["מדד", "ערך"])
        _xl_write_headers(ws, ["מדד", "ערך"])
        # remove duplicate header row written by append
        ws.delete_rows(ws.max_row - 1)
        for label, val in rows_data:
            ws.append([label, val])
        ws.append([])

    _xl_autofit(ws)

    # ── Sheet 2: משתמשים ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("משתמשים")
    ws2.sheet_view.rightToLeft = True

    user_rows = (await db.execute(
        select(User).where(User.is_admin == False).order_by(User.created_at.desc())
    )).scalars().all()

    prod_count_rows = (await db.execute(
        select(UserProduct.user_id, func.count().label("cnt")).group_by(UserProduct.user_id)
    )).all()
    prod_count_map = {r.user_id: r.cnt for r in prod_count_rows}

    ws2.append(["מ\"ס", "אימייל", "אימייל התראות", "מאומת", "פעיל", "חופשה", "Bounce",
                "סוג Bounce", "מוצרים", "מקס׳ מוצרים", "שיטת הרשמה",
                "תאריך הרשמה", "כניסה אחרונה"])
    _xl_write_headers(ws2, ["מ\"ס", "אימייל", "אימייל התראות", "מאומת", "פעיל", "חופשה", "Bounce",
                             "סוג Bounce", "מוצרים", "מקס׳ מוצרים", "שיטת הרשמה",
                             "תאריך הרשמה", "כניסה אחרונה"])
    ws2.delete_rows(ws2.max_row - 1)

    fmt = lambda dt: dt.strftime("%d/%m/%Y %H:%M") if dt else "—"
    for i, u in enumerate(user_rows, 1):
        ws2.append([
            i,
            u.email,
            u.notify_email or "—",
            "כן" if u.is_verified else "לא",
            "כן" if u.is_active else "לא",
            "כן" if u.vacation_mode else "לא",
            "כן" if u.notify_email_bounced else "לא",
            u.notify_email_bounce_type or "—",
            prod_count_map.get(u.id, 0),
            u.max_products or "—",
            "Google" if u.google_id else "אימייל",
            fmt(u.created_at),
            fmt(u.last_login_at),
        ])
    _xl_autofit(ws2)

    # ── Sheet 3: מוצרים ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("מוצרים")
    ws3.sheet_view.rightToLeft = True

    product_rows = (await db.execute(
        select(Product, func.count(UserProduct.user_id).label("watchers"))
        .outerjoin(UserProduct, UserProduct.product_id == Product.id)
        .group_by(Product.id)
        .order_by(func.count(UserProduct.user_id).desc())
    )).all()

    ws3.append(["מ\"ס", "ASIN", "שם מוצר", "סטטוס", "עוקבים", "שגיאות ברצף", "בדיקה אחרונה", "URL"])
    _xl_write_headers(ws3, ["מ\"ס", "ASIN", "שם מוצר", "סטטוס", "עוקבים", "שגיאות ברצף", "בדיקה אחרונה", "URL"])
    ws3.delete_rows(ws3.max_row - 1)

    for i, (p, watchers) in enumerate(product_rows, 1):
        ws3.append([i, p.asin, p.name or "—", p.last_status or "—", watchers,
                    p.consecutive_errors, fmt(p.last_checked), p.url or "—"])
    _xl_autofit(ws3)

    # ── Sheet 4: ביצועי מיילים ──────────────────────────────────────────────
    ws4 = wb.create_sheet("ביצועי מיילים")
    ws4.sheet_view.rightToLeft = True

    tpl_send_rows = (await db.execute(
        select(EmailSendLog.template_id, EmailSendLog.template_name,
               func.sum(EmailSendLog.sent_count).label("sent"),
               func.sum(EmailSendLog.failed_count).label("failed"))
        .group_by(EmailSendLog.template_id, EmailSendLog.template_name)
        .order_by(func.sum(EmailSendLog.sent_count).desc())
    )).all()

    open_by_tid = {r.template_id: r.opens for r in (await db.execute(
        select(EmailOpen.template_id, func.count().label("opens")).group_by(EmailOpen.template_id)
    )).all()}

    ws4.append(["תבנית", "נשלח", "נכשל", "נפתח", "Open Rate %"])
    _xl_write_headers(ws4, ["תבנית", "נשלח", "נכשל", "נפתח", "Open Rate %"])
    ws4.delete_rows(ws4.max_row - 1)

    for r in tpl_send_rows:
        sent = r.sent or 0
        opens = open_by_tid.get(r.template_id, 0)
        rate = f"{round(opens / sent * 100, 1)}%" if sent else "—"
        ws4.append([r.template_name, sent, r.failed or 0, opens, rate])
    _xl_autofit(ws4)

    # ── Sheet 5: רישומים ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("רישומים לאורך זמן")
    ws5.sheet_view.rightToLeft = True

    reg_rows = (await db.execute(
        select(cast(User.created_at, Date).label("date"), func.count().label("count"))
        .where(User.is_admin == False)
        .group_by(cast(User.created_at, Date))
        .order_by(cast(User.created_at, Date).asc())
    )).all()

    ws5.append(["תאריך", "הרשמות"])
    _xl_write_headers(ws5, ["תאריך", "הרשמות"])
    ws5.delete_rows(ws5.max_row - 1)

    for r in reg_rows:
        ws5.append([str(r.date), r.count])
    _xl_autofit(ws5)

    # ── Return file ──────────────────────────────────────────────────────────
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="amzfree-report-{file_date}.xlsx"'},
    )


@router.get("/logs")
async def get_persistent_logs(
    request: Request,
    lines: int = 500,
    grep: str = "",
):
    """Return last N lines from the persistent log file on the Railway volume.
    Auth: X-Admin-Token: ADMIN_SECRET_TOKEN  (for CLI/scripts)
       OR standard Bearer JWT admin session  (for admin panel).
    Optional ?grep= filter (case-insensitive substring match).
    """
    import os, collections
    secret = os.environ.get("ADMIN_SECRET_TOKEN", "")
    token_header = request.headers.get("x-admin-token", "")
    if not (secret and token_header == secret):
        # Fall back to JWT admin check via jose
        from jose import JWTError, jwt as jose_jwt
        from backend.auth import SECRET_KEY, ALGORITHM
        from backend.models import User as UserModel
        from backend.database import AsyncSessionLocal
        auth_header = request.headers.get("authorization", "")
        if not auth_header.startswith("Bearer "):
            raise HTTPException(status_code=401, detail="Not authenticated")
        try:
            payload = jose_jwt.decode(auth_header[7:], SECRET_KEY, algorithms=[ALGORITHM])
            user_id = int(payload.get("sub", 0))
        except (JWTError, ValueError):
            raise HTTPException(status_code=401, detail="Invalid token")
        async with AsyncSessionLocal() as db:
            user = (await db.execute(
                select(UserModel).where(UserModel.id == user_id, UserModel.is_active == True)
            )).scalar_one_or_none()
        if not user or not user.is_admin:
            raise HTTPException(status_code=403, detail="Admin access required")

    log_path = os.path.join(
        os.environ.get("BROWSER_PROFILE_DIR", "/app/browser_profile"), "logs", "app.log"
    )
    if not os.path.exists(log_path):
        return {"lines": [], "log_path": log_path, "error": "Log file not found"}

    grep_lower = grep.lower()
    buf = collections.deque(maxlen=lines)
    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\n")
            if not grep_lower or grep_lower in line.lower():
                buf.append(line)

    return {"lines": list(buf), "total_matched": len(buf), "log_path": log_path}


@router.get("/query/notifications")
async def query_notifications(
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    asin: str = "",
    email: str = "",
    limit: int = 200,
):
    """Query notification_log by ASIN and/or email. Auth: X-Admin-Token header."""
    import os
    secret = os.environ.get("ADMIN_SECRET_TOKEN", "")
    token_header = request.headers.get("x-admin-token", "")
    if not (secret and token_header == secret):
        raise HTTPException(status_code=401, detail="Not authenticated")

    stmt = (
        select(NotificationLog, User.email.label("user_email"), Product.asin.label("product_asin"), Product.name.label("product_name"))
        .join(User, NotificationLog.user_id == User.id)
        .join(Product, NotificationLog.product_id == Product.id)
        .order_by(NotificationLog.sent_at.desc())
        .limit(limit)
    )
    if asin:
        stmt = stmt.where(Product.asin == asin)
    if email:
        stmt = stmt.where(User.email == email)

    rows = (await db.execute(stmt)).all()
    return {
        "total": len(rows),
        "rows": [
            {
                "id": log.id,
                "sent_at": log.sent_at.isoformat(),
                "user_email": user_email,
                "email_to": log.email_to,
                "asin": product_asin,
                "product_name": product_name,
                "status": log.status,
                "success": log.success,
                "error_msg": log.error_msg,
            }
            for log, user_email, product_asin, product_name in rows
        ],
    }


@router.post("/check-asin")
async def check_asin_now(
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    asin: str = "",
):
    """Force-check a product by ASIN. Auth: X-Admin-Token header."""
    import os
    secret = os.environ.get("ADMIN_SECRET_TOKEN", "")
    token_header = request.headers.get("x-admin-token", "")
    if not (secret and token_header == secret):
        raise HTTPException(status_code=401, detail="Not authenticated")

    if not asin:
        raise HTTPException(status_code=400, detail="asin query param required")

    result = await db.execute(select(Product).where(Product.asin == asin))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail=f"ASIN {asin} not found in DB")

    from backend.scheduler import check_single_product
    import asyncio
    asyncio.create_task(check_single_product(product.asin, product.url))
    return {"message": f"Check triggered for {asin}", "product_id": product.id}


@router.get("/screenshot/{filename}")
async def serve_screenshot(filename: str):
    """Serve a buybox snapshot file (.html). No auth — filenames are non-guessable (ASIN+timestamp)."""
    from backend.checker import BROWSER_PROFILE_DIR
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    path = os.path.join(BROWSER_PROFILE_DIR, "screenshots", filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Snapshot not found")
    media_type = "text/html" if filename.endswith(".html") else "image/png"
    return FileResponse(path, media_type=media_type)


def _parse_price(raw: str | None) -> float:
    import re
    try:
        return float(re.sub(r"[^\d.]", "", str(raw or "")))
    except (ValueError, TypeError):
        return 0


@router.get("/blog-candidates")
async def get_blog_candidates(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    published_result = await db.execute(select(BlogPublishedAsin.asin))
    published_asins = {row[0] for row in published_result.all()}

    dismissed_result = await db.execute(select(BlogDismissedAsin.asin))
    dismissed_asins = {row[0] for row in dismissed_result.all()}

    drafts_result = await db.execute(select(BlogDraft.asin, BlogDraft.slug, BlogDraft.title))
    drafts_map = {row[0]: {"slug": row[1], "title": row[2]} for row in drafts_result.all()}

    from datetime import timezone
    cutoff = datetime.now(timezone.utc) - timedelta(hours=26)
    result = await db.execute(
        select(Product)
        .where(Product.last_status == "FREE")
        .where(
            or_(
                Product.source == "scanner",
                Product.last_checked >= cutoff,
            )
        )
    )
    products = result.scalars().all()

    # Hebrew category labels from the DB translation table (single source of truth,
    # shared with the public free-products endpoint). Falls back to English if missing.
    unique_cats = {p.amazon_category for p in products if p.amazon_category}
    cat_he_map: dict[str, str] = {}
    if unique_cats:
        trans_rows = (await db.execute(
            select(CategoryTranslation).where(CategoryTranslation.english_name.in_(unique_cats))
        )).scalars().all()
        cat_he_map = {row.english_name: row.hebrew_name for row in trans_rows}

    candidates = []
    for p in products:
        if p.asin in published_asins or p.asin in dismissed_asins:
            continue
        price = _parse_price(p.last_price)
        draft_info = drafts_map.get(p.asin)
        candidates.append({
            "asin": p.asin,
            "name": p.name or "",
            "name_he": p.name_he or "",
            "last_price": p.last_price or "",
            "price_ils": price,
            "amazon_category": p.amazon_category or "",
            "category_he": cat_he_map.get(p.amazon_category, "") if p.amazon_category else "",
            "image_url": p.image_url or "",
            "last_status": p.last_status or "",
            "url": p.url or f"https://www.amazon.com/dp/{p.asin}",
            "has_draft": draft_info is not None,
            "slug": draft_info["slug"] if draft_info else None,
        })

    candidates.sort(key=lambda x: x["price_ils"], reverse=True)
    return {"candidates": candidates, "published_count": len(published_asins)}


@router.post("/blog-candidates/{asin}/mark-published")
async def mark_blog_asin_published(
    asin: str,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    existing = await db.execute(select(BlogPublishedAsin).where(BlogPublishedAsin.asin == asin))
    if existing.scalar_one_or_none():
        return {"message": "already marked"}
    db.add(BlogPublishedAsin(asin=asin))
    await db.commit()
    return {"message": "marked"}


@router.delete("/blog-candidates/{asin}/mark-published")
async def unmark_blog_asin_published(
    asin: str,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    await db.execute(delete(BlogPublishedAsin).where(BlogPublishedAsin.asin == asin))
    await db.commit()
    return {"message": "unmarked"}


@router.post("/blog-candidates/{asin}/dismiss")
async def dismiss_blog_candidate(
    asin: str,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    existing = await db.execute(select(BlogDismissedAsin).where(BlogDismissedAsin.asin == asin))
    if not existing.scalar_one_or_none():
        db.add(BlogDismissedAsin(asin=asin))
        await db.commit()
    return {"message": "dismissed"}


class BulkDismissRequest(BaseModel):
    asins: list[str]
    dismissed: bool = True


@router.post("/blog-candidates/dismiss-bulk")
async def bulk_dismiss_blog_candidates(
    body: BulkDismissRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Dismiss (or restore) a whole filtered list of candidates in one round-trip."""
    asins = [a for a in dict.fromkeys(body.asins) if a]
    if not asins:
        return {"message": "noop", "count": 0}

    if body.dismissed:
        existing = await db.execute(
            select(BlogDismissedAsin.asin).where(BlogDismissedAsin.asin.in_(asins))
        )
        already = {row[0] for row in existing.all()}
        for asin in asins:
            if asin not in already:
                db.add(BlogDismissedAsin(asin=asin))
    else:
        await db.execute(delete(BlogDismissedAsin).where(BlogDismissedAsin.asin.in_(asins)))

    await db.commit()
    return {"message": "dismissed" if body.dismissed else "undismissed", "count": len(asins)}


async def get_product_view_counts(asins: list[str]) -> dict[str, int]:
    """Read per-ASIN page-view counters written by the website (Upstash Redis)."""
    url = os.environ.get("UPSTASH_REDIS_REST_URL")
    token = os.environ.get("UPSTASH_REDIS_REST_TOKEN")
    if not url or not token or not asins:
        return {}

    import httpx

    commands = [["get", f"product_views:{asin}"] for asin in asins]
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.post(
                f"{url}/pipeline",
                headers={"Authorization": f"Bearer {token}"},
                json=commands,
            )
        results = resp.json()
    except (httpx.HTTPError, ValueError):
        return {}

    counts = {}
    for asin, entry in zip(asins, results):
        value = entry.get("result") if isinstance(entry, dict) else None
        counts[asin] = int(value) if value else 0
    return counts


@router.get("/blog-published")
async def get_blog_published(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    result = await db.execute(select(BlogPublishedAsin).order_by(BlogPublishedAsin.marked_at.desc()))
    rows = result.scalars().all()
    view_counts = await get_product_view_counts([r.asin for r in rows])
    return {
        "published": [
            {
                "asin": r.asin,
                "slug": r.slug,
                "title": r.title,
                "marked_at": r.marked_at.isoformat() if r.marked_at else None,
                "telegram_sent_at": r.telegram_sent_at.isoformat() if r.telegram_sent_at else None,
                "facebook_sent_at": r.facebook_sent_at.isoformat() if r.facebook_sent_at else None,
                "views": view_counts.get(r.asin, 0),
            }
            for r in rows
        ]
    }


@router.post("/blog-published/{asin}/unpublish")
async def unpublish_blog_post(
    asin: str,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Remove a published post from the blog entirely: delete the HTML file (→404),
    remove its card from prices.html, drop the DB row, and dismiss the ASIN so it
    won't return as a candidate."""
    pub_row = (
        await db.execute(select(BlogPublishedAsin).where(BlogPublishedAsin.asin == asin))
    ).scalar_one_or_none()
    if not pub_row:
        raise HTTPException(404, "לא נמצא פוסט מפורסם עבור ASIN זה")

    slug = pub_row.slug
    deleted_file = False
    removed_card = False
    if slug:
        try:
            deleted_file = await delete_github_file(f"blog/{slug}.html", f"blog: unpublish {slug}")
        except Exception as e:
            logger.error("unpublish delete file error for %s: %s", slug, e, exc_info=True)
            raise HTTPException(502, f"GitHub delete error: {e}")
        try:
            removed_card = await remove_from_prices_page(slug)
        except Exception as e:
            logger.error("unpublish prices.html error for %s: %s", slug, e, exc_info=True)
            raise HTTPException(502, f"prices.html update error: {e}")

    # Drop the published row
    await db.execute(delete(BlogPublishedAsin).where(BlogPublishedAsin.asin == asin))

    # Dismiss the ASIN so it won't resurface as a blog candidate
    existing_dismiss = await db.execute(
        select(BlogDismissedAsin).where(BlogDismissedAsin.asin == asin)
    )
    if not existing_dismiss.scalar_one_or_none():
        db.add(BlogDismissedAsin(asin=asin))

    # Remove any pending (unsent) social queue entry
    try:
        from backend.models import BlogSocialQueue
        await db.execute(
            delete(BlogSocialQueue).where(
                BlogSocialQueue.asin == asin,
                BlogSocialQueue.telegram_sent.is_(False),
                BlogSocialQueue.facebook_sent.is_(False),
            )
        )
    except Exception as e:
        logger.warning("unpublish social-queue cleanup error for %s: %s", asin, e)

    await db.commit()
    return {
        "message": "unpublished",
        "asin": asin,
        "slug": slug,
        "file_deleted": deleted_file,
        "card_removed": removed_card,
    }


class EditBlogPostRequest(BaseModel):
    asin: str
    find: str
    replace: str
    replace_all: bool = False


@router.post("/blog-edit")
async def edit_blog_post(
    body: EditBlogPostRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Light manual find & replace on a published post's HTML — no Claude regen."""
    asin = body.asin.strip().upper()
    if not body.find:
        raise HTTPException(400, "יש להזין טקסט לחיפוש")

    pub_row = (await db.execute(
        select(BlogPublishedAsin).where(BlogPublishedAsin.asin == asin)
    )).scalar_one_or_none()
    if not pub_row or not pub_row.slug:
        raise HTTPException(404, "הפוסט לא נמצא")

    slug = pub_row.slug
    path = f"blog/{slug}.html"

    try:
        current, sha = await get_github_file(path)
    except Exception as e:
        logger.error("blog-edit fetch error for %s: %s", slug, e, exc_info=True)
        raise HTTPException(502, f"GitHub fetch error: {e}")

    count = current.count(body.find)
    if count == 0:
        raise HTTPException(404, {"message": "הטקסט לא נמצא בפוסט"})
    if count > 1 and not body.replace_all:
        raise HTTPException(422, {"multiple": True, "count": count,
                                  "message": f"הטקסט מופיע {count} פעמים — הוסף עוד מילים כדי שיהיה ייחודי, או החלף את כולם"})

    updated = current.replace(body.find, body.replace)

    try:
        await commit_to_github(path, updated, f"blog: edit {slug}", sha=sha)
    except Exception as e:
        logger.error("blog-edit commit error for %s: %s", slug, e, exc_info=True)
        raise HTTPException(502, f"GitHub commit error: {e}")

    return {
        "slug": slug,
        "replacements": count,
        "preview_url": f"https://www.amzfreeil.com/blog/{slug}.html",
    }


@router.get("/blog-social-queue")
async def get_blog_social_queue(
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    from backend.models import BlogSocialQueue
    sent_result = await db.execute(
        select(BlogSocialQueue).where(
            BlogSocialQueue.telegram_sent.is_(True), BlogSocialQueue.facebook_sent.is_(True)
        )
    )
    for row in sent_result.scalars().all():
        await db.delete(row)
    await db.commit()

    result = await db.execute(select(BlogSocialQueue).order_by(BlogSocialQueue.scheduled_at.asc()))
    rows = result.scalars().all()
    return {
        "queue": [
            {
                "id": r.id,
                "asin": r.asin,
                "kind": r.kind or "review",
                "slug": r.slug,
                "title": r.title,
                "scheduled_at": r.scheduled_at.isoformat() if r.scheduled_at else None,
                "manual": bool(r.manual),
                "telegram_sent": r.telegram_sent,
                "facebook_sent": r.facebook_sent,
            }
            for r in rows
        ]
    }


class RescheduleBlogSocialRequest(BaseModel):
    scheduled_at: str


@router.patch("/blog-social-queue/{row_id}/schedule")
async def reschedule_blog_social_post(
    row_id: int,
    body: RescheduleBlogSocialRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Set the broadcast time of a queued post by hand. Times outside the
    06:00-22:00 window (or too close to another post) are allowed but warned about.

    Keyed by row id rather than ASIN — editorial guides ride the same queue with
    asin=NULL."""
    from backend.models import BlogSocialQueue
    from backend.scheduler import blog_social_time_warnings, parse_manual_blog_social_time

    row = (
        await db.execute(select(BlogSocialQueue).where(BlogSocialQueue.id == row_id))
    ).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "הפוסט לא נמצא בתור השידור")

    try:
        scheduled_at = parse_manual_blog_social_time(body.scheduled_at)
    except ValueError as e:
        raise HTTPException(400, str(e))

    others = (await db.execute(
        select(BlogSocialQueue.scheduled_at).where(BlogSocialQueue.id != row_id)
    )).scalars().all()

    row.scheduled_at = scheduled_at
    row.manual = True
    await db.commit()

    logger.info(f"[blog_social_queue] admin rescheduled {row.slug} to {scheduled_at.isoformat()}")
    return {
        "scheduled_at": scheduled_at.isoformat(),
        "warnings": blog_social_time_warnings(scheduled_at, list(others)),
    }


@router.delete("/blog-candidates/{asin}/dismiss")
async def undismiss_blog_candidate(
    asin: str,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    await db.execute(delete(BlogDismissedAsin).where(BlogDismissedAsin.asin == asin))
    await db.commit()
    return {"message": "undismissed"}


@router.get("/blog-draft/{asin}")
async def get_blog_draft(
    asin: str,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Look up a single draft by ASIN.

    Used by the admin UI to recover from a lost response: draft generation takes
    45-70s and the connection sometimes drops after the work already succeeded,
    which used to show a false "failed" and tempt a retry that re-runs Claude and
    re-commits to GitHub.
    """
    row = (await db.execute(
        select(BlogDraft).where(BlogDraft.asin == asin.strip().upper())
    )).scalar_one_or_none()
    if not row:
        return {"exists": False}
    repo = os.getenv("GITHUB_REPO", "")
    return {
        "exists": True,
        "slug": row.slug,
        "title": row.title or "",
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "github_url": f"https://github.com/{repo}/blob/main/blog/{row.slug}.html",
        "preview_url": f"https://www.amzfreeil.com/blog/{row.slug}.html",
    }


async def _run_blog_draft(body: GenerateBlogDraftRequest, db: AsyncSession) -> dict:
    """Fetch → generate → build → commit for one ASIN.

    Shared by the synchronous single-draft endpoint and the batch worker, so both
    paths stay byte-identical in what they produce. Raises HTTPException; the
    batch worker turns those into a `failed` job row.
    """
    asin = body.asin.strip().upper()

    if body.manual_title and body.manual_features is not None:
        # Manual override — user provided data directly (e.g. when API is blocked)
        partner_tag = os.getenv("AMAZON_AFFILIATE_TAG") or os.getenv("AMAZON_PARTNER_TAG", "amzfreeil-20")
        product = {
            "asin": asin,
            "title": body.manual_title.strip(),
            "features": [f.strip() for f in body.manual_features if f.strip()],
            # The ASIN-guess pattern (images/P/{asin}.01.L.jpg) returns a blank 1x1
            # placeholder for many newer products — so prefer a real URL if provided.
            "image": (body.manual_image or "").strip()
                     or f"https://images-na.ssl-images-amazon.com/images/P/{asin}.01.L.jpg",
            "model": "",
            "brand": (body.manual_brand or "").strip(),
            "category": (body.manual_category or "").strip(),
            "url": f"https://www.amazon.com/dp/{asin}?tag={partner_tag}",
        }
    else:
        try:
            product = await fetch_amazon_product(asin)
        except ValueError as e:
            if "ItemNotAccessible" in str(e):
                logger.info("blog-draft: ItemNotAccessible for %s — asking for manual input", asin)
                raise HTTPException(
                    status_code=422,
                    detail={"blocked": True, "asin": asin, "message": f"המוצר {asin} חסום ב-Amazon API — הזן title ו-features ידנית"},
                )
            logger.error("blog-draft Amazon API error for %s: %s", asin, e, exc_info=True)
            raise HTTPException(502, f"Amazon API error: {e}")
        except Exception as e:
            logger.error("blog-draft Amazon API error for %s: %s", asin, e, exc_info=True)
            raise HTTPException(502, f"Amazon API error: {e}")

        missing_brand = not product.get("brand")
        missing_category = not product.get("category")
        if missing_brand or missing_category:
            if body.manual_brand or body.manual_category:
                if missing_brand:
                    product["brand"] = (body.manual_brand or "").strip()
                if missing_category:
                    product["category"] = (body.manual_category or "").strip()
            else:
                logger.info("blog-draft: missing brand/category for %s — asking for manual input", asin)
                raise HTTPException(
                    status_code=422,
                    detail={
                        "missing_fields": True,
                        "asin": asin,
                        "missing_brand": missing_brand,
                        "missing_category": missing_category,
                        "message": f"למוצר {asin} חסרים מותג/קטגוריה ב-Amazon API — הזן ידנית",
                    },
                )

    try:
        content = await generate_with_claude(product, body.israel_price, body.amazon_price, body.min_order_49)
    except Exception as e:
        logger.error("blog-draft Claude API error for %s: %s", asin, e, exc_info=True)
        raise HTTPException(502, f"Claude API error: {e}")

    # If this ASIN is already published, update the live post in place: reuse
    # the original slug (Claude invents a new one each run) so we overwrite the
    # same file instead of forking a new URL.
    pub_row = (await db.execute(
        select(BlogPublishedAsin).where(BlogPublishedAsin.asin == asin)
    )).scalar_one_or_none()
    if pub_row and pub_row.slug:
        content["slug"] = pub_row.slug

    # Same story for a draft that hasn't been published yet: without this, every
    # regeneration minted a fresh slug and left blog/{old-slug}.html orphaned on
    # GitHub — noindex, unlinked, never cleaned up. Reuse the draft's own slug so
    # the same file is overwritten instead.
    draft_row = (await db.execute(
        select(BlogDraft).where(BlogDraft.asin == asin)
    )).scalar_one_or_none()
    if draft_row and draft_row.slug and not pub_row:
        content["slug"] = draft_row.slug

    try:
        html = build_post_html(product, content, body.israel_price, body.amazon_price, body.min_order_49, body.voltage_warning)
    except Exception as e:
        logger.error("blog-draft build_post_html error for %s: %s", asin, e, exc_info=True)
        raise HTTPException(502, f"Build HTML error: {e}")

    slug = content["slug"]

    # Already-published post: strip the noindex meta build_post_html adds *before*
    # the commit, so a single write keeps the post indexed. (Committing then
    # calling publish_draft did two writes to the same file and raced GitHub's
    # eventual consistency → 409 Conflict.)
    commit_message = f"blog: draft {content.get('title_short', slug)}"
    if pub_row:
        html = html.replace(
            '<meta name="robots" content="noindex,nofollow" />\n', ""
        ).replace(
            '<meta name="robots" content="noindex,nofollow" />', ""
        )
        commit_message = f"blog: update {content.get('title_short', slug)}"

    try:
        await commit_to_github(
            path=f"blog/{slug}.html",
            content=html,
            message=commit_message,
        )
    except Exception as e:
        logger.error("blog-draft GitHub commit error for %s: %s", asin, e, exc_info=True)
        raise HTTPException(502, f"GitHub commit error: {e}")

    # Already-published post: refresh the stored title and return without creating
    # a new draft row (no re-publish side effects like social queue).
    if pub_row:
        pub_row.title = content.get("title_short") or content.get("title_he", "")
        await db.commit()
        repo = os.getenv("GITHUB_REPO", "")
        return {
            "republished": True,
            "slug": slug,
            "title": content.get("title_he", ""),
            "github_url": f"https://github.com/{repo}/blob/main/blog/{slug}.html",
            "preview_url": f"https://www.amzfreeil.com/blog/{slug}.html",
        }

    # draft_row was already looked up above, to pin the slug before the build.
    if draft_row:
        draft_row.slug = slug
        draft_row.title = content.get("title_he", "")
        draft_row.title_short = content.get("title_short", "")
        draft_row.israel_price = body.israel_price
        draft_row.amazon_price = body.amazon_price
        draft_row.image_url = product.get("image", "")
        draft_row.min_order_49 = body.min_order_49
        draft_row.voltage_warning = body.voltage_warning
    else:
        db.add(BlogDraft(
            asin=asin,
            slug=slug,
            title=content.get("title_he", ""),
            title_short=content.get("title_short", ""),
            israel_price=body.israel_price,
            amazon_price=body.amazon_price,
            image_url=product.get("image", ""),
            min_order_49=body.min_order_49,
            voltage_warning=body.voltage_warning,
        ))
    await db.commit()

    repo = os.getenv("GITHUB_REPO", "")
    return {
        "slug": slug,
        "title": content.get("title_he", ""),
        "github_url": f"https://github.com/{repo}/blob/main/blog/{slug}.html",
        "preview_url": f"https://www.amzfreeil.com/blog/{slug}.html",
    }


@router.post("/blog-draft")
async def generate_blog_draft(
    body: GenerateBlogDraftRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    return await _run_blog_draft(body, db)


class BatchBlogDraftRequest(BaseModel):
    items: list[GenerateBlogDraftRequest]


# A stuck `running` row means the container restarted mid-draft (Railway kills
# in-flight asyncio tasks); nothing will ever finish it, so surface it as failed.
BLOG_DRAFT_JOB_TIMEOUT_MIN = 15


async def _run_batch(batch_id: str) -> None:
    """Background worker: generate every pending draft in a batch, N at a time.

    Drafts are independent — each commits its own blog/{slug}.html — so the only
    reason to cap concurrency is upstream rate limits (Anthropic, Amazon PA-API).
    """
    concurrency = int(os.getenv("BLOG_DRAFT_CONCURRENCY", "3"))
    sem = asyncio.Semaphore(concurrency)

    async with AsyncSessionLocal() as db:
        jobs = (await db.execute(
            select(BlogDraftJob).where(BlogDraftJob.batch_id == batch_id)
        )).scalars().all()
        job_ids = [j.id for j in jobs]

    async def run_one(job_id: int) -> None:
        async with sem:
            # A session per job: a failure rolls back only its own row, and the
            # concurrent drafts don't share a connection.
            async with AsyncSessionLocal() as db:
                job = await db.get(BlogDraftJob, job_id)
                if not job or job.status != "pending":
                    return
                job.status = "running"
                await db.commit()

                body = GenerateBlogDraftRequest(
                    asin=job.asin,
                    israel_price=job.israel_price,
                    amazon_price=job.amazon_price,
                    min_order_49=job.min_order_49,
                    voltage_warning=job.voltage_warning,
                )
                try:
                    result = await _run_blog_draft(body, db)
                    job.status = "done"
                    job.slug = result.get("slug")
                    job.title = result.get("title") or ""
                except HTTPException as e:
                    detail = e.detail
                    if isinstance(detail, dict):
                        detail = detail.get("message") or str(detail)
                    # Drop whatever the aborted pipeline left pending, so only the
                    # job's own failure status gets committed below.
                    await db.rollback()
                    job = await db.get(BlogDraftJob, job_id)
                    job.status = "failed"
                    job.error = str(detail)[:1000]
                    logger.error("batch draft failed for %s: %s", job.asin, detail)
                except Exception as e:
                    await db.rollback()
                    job = await db.get(BlogDraftJob, job_id)
                    job.status = "failed"
                    job.error = str(e)[:1000]
                    logger.error("batch draft failed for %s: %s", job.asin, e, exc_info=True)

                job.finished_at = datetime.utcnow()
                await db.commit()

    await asyncio.gather(*(run_one(jid) for jid in job_ids), return_exceptions=True)
    logger.info("batch %s finished (%d jobs)", batch_id, len(job_ids))


@router.post("/blog-draft/batch")
async def generate_blog_drafts_batch(
    body: BatchBlogDraftRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Queue N drafts and return immediately — the admin polls for status."""
    if not body.items:
        raise HTTPException(400, "no items")

    batch_id = str(uuid.uuid4())
    for item in body.items:
        db.add(BlogDraftJob(
            batch_id=batch_id,
            asin=item.asin.strip().upper(),
            israel_price=item.israel_price,
            amazon_price=item.amazon_price,
            min_order_49=item.min_order_49,
            voltage_warning=item.voltage_warning,
        ))
    await db.commit()

    asyncio.create_task(_run_batch(batch_id))
    return {"batch_id": batch_id, "count": len(body.items)}


@router.get("/blog-draft/batch/{batch_id}")
async def get_blog_draft_batch(
    batch_id: str,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    jobs = (await db.execute(
        select(BlogDraftJob).where(BlogDraftJob.batch_id == batch_id).order_by(BlogDraftJob.id)
    )).scalars().all()
    if not jobs:
        raise HTTPException(404, "batch not found")

    cutoff = datetime.utcnow() - timedelta(minutes=BLOG_DRAFT_JOB_TIMEOUT_MIN)
    stale = False
    for job in jobs:
        if job.status == "running" and job.created_at and job.created_at.replace(tzinfo=None) < cutoff:
            job.status = "failed"
            job.error = "העבודה נקטעה (ככל הנראה אתחול שרת) — נסה שוב"
            job.finished_at = datetime.utcnow()
            stale = True
    if stale:
        await db.commit()

    repo = os.getenv("GITHUB_REPO", "")
    return {
        "batch_id": batch_id,
        "done": all(j.status in ("done", "failed") for j in jobs),
        "items": [{
            "asin": j.asin,
            "status": j.status,
            "error": j.error,
            "slug": j.slug,
            "title": j.title,
            "github_url": f"https://github.com/{repo}/blob/main/blog/{j.slug}.html" if j.slug else None,
            "preview_url": f"https://www.amzfreeil.com/blog/{j.slug}.html" if j.slug else None,
        } for j in jobs],
    }


class PublishBlogDraftRequest(BaseModel):
    asin: str
    slug: str
    # Optional admin-chosen broadcast time ("2026-07-20T09:30", IL time).
    # Empty/None = the scheduler draws a random slot as usual.
    scheduled_at: str | None = None


@router.post("/blog-publish")
async def publish_blog_draft(
    body: PublishBlogDraftRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    asin = body.asin
    slug = body.slug

    draft_row = (await db.execute(select(BlogDraft).where(BlogDraft.asin == asin))).scalar_one_or_none()

    # Validate the manual time before touching GitHub, so a bad time can't leave
    # a post published with no queue entry.
    manual_at = None
    if body.scheduled_at:
        from backend.scheduler import parse_manual_blog_social_time
        try:
            manual_at = parse_manual_blog_social_time(body.scheduled_at)
        except ValueError as e:
            raise HTTPException(400, str(e))

    try:
        await publish_draft(slug)
    except Exception as e:
        raise HTTPException(502, f"GitHub publish error: {e}")

    if draft_row and draft_row.amazon_price:
        try:
            await add_to_prices_page(
                asin=asin,
                slug=slug,
                title_short=draft_row.title_short or draft_row.title,
                israel_price=draft_row.israel_price,
                amazon_price=draft_row.amazon_price,
                image_url=draft_row.image_url or "",
            )
        except Exception as e:
            raise HTTPException(502, f"prices.html update error: {e}")

    scheduled_at = None
    social_warnings: list[str] = []
    if draft_row:
        title = draft_row.title_short or draft_row.title
        try:
            scheduled_at, social_warnings = await queue_blog_social_post(
                asin, slug, title, draft_row.image_url, draft_row.amazon_price,
                draft_row.israel_price, manual_at=manual_at,
            )
        except Exception as e:
            logger.warning(f"blog-publish queue error: {e}")

    existing_pub = await db.execute(select(BlogPublishedAsin).where(BlogPublishedAsin.asin == asin))
    pub_row = existing_pub.scalar_one_or_none()
    pub_title = (draft_row.title_short or draft_row.title) if draft_row else None
    if pub_row:
        pub_row.slug = slug
        pub_row.title = pub_title
    else:
        db.add(BlogPublishedAsin(asin=asin, slug=slug, title=pub_title))

    await db.execute(delete(BlogDraft).where(BlogDraft.asin == asin))
    await db.commit()

    repo = os.getenv("GITHUB_REPO", "")
    return {
        "slug": slug,
        "url": f"https://www.amzfreeil.com/blog/{slug}.html",
        "github_url": f"https://github.com/{repo}/blob/main/blog/{slug}.html",
        "social_scheduled_at": scheduled_at.isoformat() if scheduled_at else None,
        "social_warnings": social_warnings,
    }


class RescueDraftRequest(BaseModel):
    asin: str
    slug: str
    title: str = ""


@router.post("/blog-rescue-draft")
async def rescue_draft_from_dismissed(
    body: RescueDraftRequest,
    admin: Annotated[User, Depends(get_current_admin)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Move an ASIN from blog_dismissed_asins to blog_drafts (one-time migration)."""
    await db.execute(delete(BlogDismissedAsin).where(BlogDismissedAsin.asin == body.asin))
    existing = await db.execute(select(BlogDraft).where(BlogDraft.asin == body.asin))
    if not existing.scalar_one_or_none():
        db.add(BlogDraft(asin=body.asin, slug=body.slug, title=body.title))
    await db.commit()
    return {"message": "rescued", "asin": body.asin, "slug": body.slug}
